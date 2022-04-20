import discord
from discord.ext import commands
import asyncio
import json
import requests
from bs4 import BeautifulSoup
import openpyxl


client = discord.Client()
token = 'OTMwMDYxMTQ5NDQ0MTQxMDU2.YdwY8Q.z6_DeVMfyEplskhQi5hsMproS7E'
bot = commands.Bot(command_prefix='!')


@client.event
async def on_ready():
    print('Logged in as')
    print(client.user.name)
    print(client.user.id)
    print('------')


    while True:
        openChannelIDlist = openpyxl.load_workbook("./channel_id.xlsx")
        wb = openChannelIDlist.active

        for i in range(1, 1000):
            if "_" not in wb["B" + str(i)].value:
                channel = client.get_channel(int(wb["B" + str(i)].value))
                await channel.send(wb["A1"].value)

            elif wb["B" + str(i)].value == "_":
                pass

        openChannelIDlist.save("channel_id.xlsx")
        await asyncio.sleep(14400)


@client.event
async def on_message(message):
    if "시작시간" in message.content:
        ch = client.get_channel(875723288121335850)
        await ch.send("@here {}님이 {}에서 파티구인중입니다! 많참부!".format(message.author.mention, message.channel.mention))

    if message.content == "성복아 안녕":
        await message.channel.send("{} 반갑습니다! 선생님!".format(message.author.mention))

    if message.content == "성복 아쎄이!":
        await message.channel.send("악! {}해병님 안녕하십니까!".format(message.author.mention))

    if message.content == "성복이 기열!":
        await message.channel.send("악!")
        
    if message.content == "성복이 입수!":
        await message.channel.send("악! 역돌격 실시!")

    if message.content == "성복이 좋아?":
        await message.channel.send("끼잉낑..")

    if message.content == "성복아 물":
        await message.channel.send("악! 따듯한물을 드려야할지 차가운물을 드려야할지 허락을 구하는 것을 여쭤봐도 괜찮으실지를 물어봐도 되겠습니까?")

    if message.content == "따뜻한 물 줘":
        await message.channel.send("악! 따뜻한 물한잔 대접해드리겠습니다")

    if message.content == "차가운 물 줘":
        await message.channel.send("악! 차가운 물한잔 대접해드리겠습니다")

    if message.content == "굿밤":
        await message.channel.send("{}님 좋은밤 되세요!".format(message.author.mention))

    if message.content == "굿모닝":
        await message.channel.send("{}님 좋은 아침입니다!".format(message.author.mention))

    if message.content == "굿모링":
        await message.channel.send("{}님 좋은 아침입니다!".format(message.author.mention))

    if "라인업 구상" in message.content:
        ch = client.get_channel(875723288121335850)
        await ch.send("@here {}님이 {}에서 내전인원 구인중입니다! 많참부!")




    def check(m):
        return m.author == message.author and m.channel == message.channel

    if message.content.startswith("?공지"):
        if message.author.id == 525326347745361920 or message.author.id == 394482863459926018 or message.author.id == 410506954151297025 or message.author.id == 336367088920690693 or message.author.id == 816403916958466088 or message.author.id == 504648898636414977:
            openChannellist = openpyxl.load_workbook("./channel_id.xlsx")
            wb = openChannellist.active

            message_content = message.content.replace("?공지 ", "")
            wb["A1"].value = message_content
            await message.channel.send(f"4시간마다 __{message_content}__ 라는 공지를 전송합니다.")
            openChannellist.save("channel_id.xlsx")

        else:
            await message.channel.send("공지를 보낼 권한이 없습니다.")

    if message.content.startswith("/알림설정"):
        channelID = str(message.channel.id)

        openChannellist = openpyxl.load_workbook("./channel_id.xlsx")
        wb = openChannellist.active

        onOffEmbed = discord.Embed(title="공지를 받으시려면 1, 받지 않으시려면 2를 입력해주세요!", color=0xFF9900)
        await message.channel.send(embed=onOffEmbed)

        try:
            onOff = await client.wait_for("message", check=check, timeout=60)

            if onOff.content == "1":
                for i in range(1, 1000):
                    if wb["B" + str(i)].value == channelID:
                        await message.channel.send("해당 채널은 이미 알림을 받는 채널입니다.")
                        break

                    elif wb["B" + str(i)].value == "_":
                        wb["B" + str(i)].value = channelID
                        await message.channel.send("해당 채널을 공지를 받는 채널로 설정했습니다.")
                        break

                openChannellist.save("channel_id.xlsx")


            elif onOff.content == "2":
                for i in range(1, 1000):
                    if wb["B" + str(i)].value == channelID:
                        wb["B" + str(i)].value = "_"
                        await message.channel.send("해당 채널을 공지를 받지 않는 채널로 설정했습니다.")
                        break

                    elif wb["B" + str(i)].value == "_":
                        await message.channel.send("이미 공지를 받지 않는 채널입니다.")
                        break

                openChannellist.save("channel_id.xlsx")

        except asyncio.exceptions.TimeoutError:
            error = discord.Embed(title="시간을 초과하셨습니다.\n60초 내로 입력해주세요.", color=0xFF9900)
            error.set_footer(text="다시 시도는 ?알림설정 을 입력해주세요.")


def find_first_channel(channels):
    position_array = [i.position for i in channels]

    for i in channels:
        if i.position == min(position_array):
            return i


@client.event
async def on_member_join(self, member):
    msg = "<@{}>님이 서버에 들어오셨어요. 환영합니다.".format(str(member.id))
    await find_first_channel(member.guild.text_channels).send(msg)
    return None


async def on_member_remove(self, member):
    msg = "<@{}>님이 서버에서 나가거나 추방되었습니다.".format(str(member.id))
    await find_first_channel(member.guild.text_channels).send(msg)
    return None


client.run(token)